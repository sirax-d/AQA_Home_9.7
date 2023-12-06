from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
import os

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
TMP_DIR = os.path.join(CURRENT_DIR, 'tmp')
RESOURSES_DIR = os.path.join(CURRENT_DIR, 'resourses')


def test_create_files_zip():
    list_files = ['test.csv', 'test.xlsx', 'test.pdf']  # список файлов которые будут добавлены в архив
    zip_name = 'files.zip'  # имя архива
    os.chdir(TMP_DIR)  # переход в директорию
    with ZipFile(zip_name, 'w') as zipbox:  # создание архива
        for file in list_files:  # добавление файлов в архив
            zipbox.write(file)
    assert os.path.exists(zip_name) is True  # проверка наличия архива


def test_change_zip_directory():
    change_directory = '..'
    os.chdir(change_directory)  # переход в директорию
    zip_name = 'files.zip'
    if not os.path.exists('resourses/'):
        os.mkdir('resourses')
    source_zip = os.path.abspath('tmp/files.zip')  # адрес архива
    destination_zip = os.path.abspath('resourses/files.zip')  # адрес куда будет перемещен архив
    os.replace(source_zip, destination_zip)  # перемещение архива
    assert os.path.exists(destination_zip) is True  # проверка наличия архива


def test_files_without_zip():
    zip_name = 'files.zip'
    with ZipFile(os.path.join(RESOURSES_DIR, zip_name), 'r') as zipbox:
        with zipbox.open('test.pdf') as pdf_file:
            pdf_reader = PdfReader(pdf_file)
            numbers_pages = len(pdf_reader.pages)
            exact_test = pdf_reader.pages[5].extract_text()
            print(exact_test)
            assert numbers_pages == 10  # проверка количества страниц в pdf файле
            assert '1.2 How towrite adocument' in exact_test  # проверка содержимого страницы
        with zipbox.open('test.xlsx') as xlsx_file:
            wb = load_workbook(xlsx_file)
            sheet = wb.active
            cell = sheet['C1']
            result_assert = (sheet.cell(row=5, column=3).value)
            assert cell.value == 'EmpName'  # проверка содержимого ячейки
            assert result_assert == 'Jacob'
        with zipbox.open('test.csv') as csv_file:
            for line_number, string in enumerate(csv_file, start=1):
                if line_number == 3:
                    assert string == b'"John ""Da Man""",Repici,120 Jefferson St.,Riverside, NJ,08075\n'  # проверка совпадения строки
                print(line_number, string)
        assert line_number == 6  # проверка количества строк в файле
